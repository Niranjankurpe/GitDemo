import openpyxl
#class SearchData:
class HomePageData:
    #test_data_homepage = ["web development"]

    @staticmethod
    def getTestData(course_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\niran\\PycharmProjects\\python_Coursera\\courseradata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == course_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]


